import os
import zipfile
import chardet
import re
import openpyxl

# function to change source of a given .docx file. Necessary for EDU consulting
# team because the main deliverables are .docx files that have links to the EIS
# model. same goes for the pdga. so this function changes the source of those
# links in the docx.


# currently the parser adding quotes around the path is fine, but causing
# problems when splitting for vizualization section. Need to figure out how to
# adapt to that particular situation.
def make_contents_dict(file_path):
    """For diagnostics, takes docx or other openXML format file, parses the
    contents into a dict with xml file names as keys and xml contents as
    values."""
    z = zipfile.ZipFile(file_path)
    contents = {}
    for name in z.namelist():
        data = z.open(name)
        read_data = data.read()
        contents[name] = read_data
        data.close()
    return(contents)

# need to add logic to ensure new is not none.


def replace_contents_dict(old, new, contents_dict):
    """Also for diagnostics, Looks at a dict from make_contents_dict and
    replaces old string with new string in every xml."""
    contents_dict_replaced = {}
    for name in contents_dict:
        contents_dict_replaced[name] = bytearray(
            contents_dict[name]).replace(old, new)
        contents_dict_replaced[name] = bytes(contents_dict_replaced[name])
    return(contents_dict_replaced)


def generate_occurance_list(contents_dict, search_string):
    """THis is for diagnostics, iterates through unpacked docx and shows the
    first index where search_string appears in each individual xml file."""
    occurance_list = {}
    for name in contents_dict:
        if contents_dict[name].find(search_string) != -1:
            occurance_list[name] = contents_dict[name].find(search_string)
    return(occurance_list)


def file_path_parser(file_path):
    """Takes a filepath, and parses into a quadruple slash path, and a double
    slash path, as those are the two formats that exist within the xml of the
    files. The body of the docx in the xml is double slashes, and the
    visulaizations are quadruple slash paths"""
    if type(file_path) == str:
        if file_path.count("\\\\") == 0:
            double_slash_path = file_path
            double_slash_path_tail = os.path.split(
                double_slash_path)[1]
            quad_slash_path = file_path.replace("\\", "\\\\")

        elif file_path.count("\\\\") > 0:
            quad_slash_path = file_path
            double_slash_path = file_path.replace("\\\\", "\\")
            double_slash_path_tail = os.path.split(
                double_slash_path)[1]

        if " " in quad_slash_path and not quad_slash_path.startswith(
                "\"") and not quad_slash_path.endswith("\""):
            quad_slash_path = "\"" + quad_slash_path + "\""

        else:
            pass
    return(double_slash_path, quad_slash_path)

# need to add logic to ensure new is not none, as that will thoroughly
# bork the reports.


def change_source(old, new, file_path):
    """Opens any openXML based file as a zipfile, then iterates through each element of the
    docx/ppt underlying openXML format, which are .xml files we read in as bytes
    objects. replaces links to excel files in each .xml file with a new path
    link."""
    # check if string, convert to bytes with ascii encoding
    if type(old) == str:
        double_old, quad_old = file_path_parser(old)

    if type(new) == str:
        double_new, quad_new = file_path_parser(new)

    if type(new) and type(old) == str or bytes or bytearray:
        # convert docx into zipfile
        zipified_docx = zipfile.ZipFile(file_path)
        # initialize dictionary to store names of files and contained data.
        subfile_name_contents = {}
        updated_xml_file_list = []

        for name in zipified_docx.namelist():
            data = zipified_docx.open(name)
            print(f"opening {name} now to read")
            read_data = data.read()
            # convert read_data into byte array, and replace the old bytes with
            # the new bytes. and old tail with new tail (for viz)
            encoding = chardet.detect(read_data)["encoding"]
            if encoding == None:
                encoding = "ascii"

            subfile_name_contents[name] = bytearray(read_data).replace(bytes(
                quad_old, encoding), bytes(quad_new, encoding)).replace(bytes(double_old, encoding),
                                                                        bytes(double_new,
                                                                              encoding))

            data.close()
            updated_xml_file_list.append(subfile_name_contents[name])

        # open docx as writeable to write our changed data into.
        writeable_zipified_docx = zipfile.ZipFile(file_path, mode='w')

        for name_mod_xml_pair in zip(zipified_docx.namelist(), updated_xml_file_list):
            file_to_write = writeable_zipified_docx.open(
                name_mod_xml_pair[0], mode='w')
            file_to_write.write(name_mod_xml_pair[1])
            print(
                f"writing {len(name_mod_xml_pair[1])} bytes to {name_mod_xml_pair[0]} now")
            file_to_write.close()

    else:
        print("Unsupported type for replacement, input either string or bytes object.")


def run_excel_links(model_path, file_path):
    # takes file_path, iterates through every link that contains text
    # fields in the xml and replaces
    # with the corresponding value from the excel workbook referenced in
    # the document.
    text_docs = ["word/document.xml", "word/footnotes.xml"]

    zipified_docx = zipfile.ZipFile(file_path)
    subfile_name_contents = {}
    openpyxl_model = openpyxl.load_workbook(model_path, data_only=True)
    for xml_doc in zipified_docx.namelist():
        if xml_doc in text_docs:
            data = zipified_docx.open(xml_doc)
            subfile_name_contents[xml_doc] = data.read()
            data.close()
            subfile_name_contents[xml_doc] = replace_xml_fields_with_data_from_model(
                subfile_name_contents[xml_doc], openpyxl_model)
    writeable_zipified_docx = zipfile.ZipFile(file_path, mode='w')
    for xml_doc in writeable_zipified_docx.namelist():
        if xml_doc in text_docs:
            file_to_write = writeable_zipified_docx.open(xml_doc, mode='w')
            file_to_write.write(subfile_name_contents[xml_doc])
            print(
                f"writing {len(subfile_name_contents[xml_doc])} bytes to {xml_doc}")
            file_to_write.close()


def parse_xml_link(link):
    """Takes an xml link to an excel file, and parses out the path to the
    workbook, the sheet the link is referencing, and the cell row and column
    values. This is important as we will pass it's results into openpyxl data
    only version of the model to pull data."""
    excel_link = re.search(b"Excel.SheetMacroEnabled\.12 (.*?) </", link)
    if excel_link:
        modelre = re.search(b"(.*?).xlsm", excel_link.group(1))
        modellen = len(modelre.group(1).split(b" "))
        sheet_cell_ref = excel_link.group(1).split(b" ")[modellen]
        sheet, cell = sheet_cell_ref.split(b"!")
        row_col_match = re.findall(b"\d+", cell)
        row_col_nums = [int(match) for match in row_col_match]
        row, col = row_col_nums
        return(sheet, row, col)
    else:
        return([None, None, None, None])


def find_all_xml_fields(xml_doc):
    formatting_list = []
    formatting_regex = re.compile(
        b'''"separate"/></w:r><w:r w:rsidR=(.*?)w:t>(.*?)</w:t></w:r>''')
    text_list = []
    text_regex = re.compile(
        b'''"separate"/></w:r><w:r w:rsidR=(.*?)w:t>(.*?)</w:t></w:r>''')
    model_data = []
    field_links = re.findall(
        b'''<w:fldChar w:fldCharType="begin"/>.*?<w:fldChar w:fldCharType="end"/>''', xml_doc)
    for link in field_links:
        if formatting_regex.search(link) is not None:
            formatting_list.append(formatting_regex.search(link))
        else:
            formatting_list.append(None)

        if text_regex.search(link) is not None:
            text_list.append(text_regex.search(link))
        else:
            text_list.append(None)

        model_data.append([parse_xml_link(link)])
    xml_field_list = list(zip(field_links, formatting_list, text_list,
                              model_data))
    return(xml_field_list)


def replace_xml_fields_with_data_from_model(xml_doc, openpyxl_model):
    encoding = chardet.detect(xml_doc)["encoding"]
    xml_doc_changed = bytearray(xml_doc)
    xml_field_list = find_all_xml_fields(xml_doc)
    print(len(xml_field_list), "number of xml fields in xml_doc")
    for field in xml_field_list:
        link_field = field[0]
        formatting = field[1]
        text = field[2]
        model_data = field[3]
        if model_data[0][0] is not None:
            sheet = model_data[0][0].replace(b'"', b'')
            worksheet = openpyxl_model[sheet.decode(encoding)]
            row, col = model_data[0][1], model_data[0][2]
            cell = worksheet.cell(row=row, column=col)
#             if formatting is not None and text is not None:
#                 replaced_field = re.sub(
#                     re.escape(text.string), bytes(str(cell.value), encoding),
#                     formatting.string)
#                 xml_doc_changed = re.sub(
#                     re.escape(link_field), replaced_field, xml_doc_changed)
#             else:
            replaced_field = f'''<w:r><w:t>{cell.value}</w:t></w:r>'''
            replaced_field = bytes(replaced_field, encoding)
            xml_doc_changed = re.sub(
                re.escape(link_field), replaced_field, xml_doc_changed)
        print(field, "/n", replaced_field)

    return(xml_doc_changed)


if __name__ == "__main__":
    file = r"C:\Users\michael.gilman\Desktop\mr1 - Copy.docx"
    old = r"C:\EIS\US_EIS.xlsm"
    new = r"G:\Shared drives\EDU Consulting - Private\Completed Reports\US\Florida\Florida Gateway\EIS\2021\Drafts\FGC_2021_EIS_Draft.xlsm"
    old_double, old_quad = file_path_parser(old)
    new_double, new_quad = file_path_parser(new)
    cd = make_contents_dict(file)
    cdnames = []
    for i in cd:
        cdnames.append(i)
    # openpyxl_model = openpyxl.load_workbook(new, data_only=True)

    change_source(old, new, file)
    run_excel_links(new, file)
