import requests
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import datetime
from json.decoder import JSONDecodeError
import platform
import sys

# attempt to make it easier to import template from google drive, but this didn't work with the load_workbook() function. Apparently this works for python file imports...
if platform.system() == 'Windows':
    sys.path.append(r"G:\\Shared drives\\EDU Consulting - Private\\PDGA report masters and surveys\\US\\PDGA Survey Template")
elif platform.system() == 'Darwin':
    sys.path.append(r"\\Google Drive\\Shared drives\\EDU Consulting - Private\\PDGA report masters and surveys\\US\\PDGA Survey Template")
else:
    print('Error: Invalid OS detected')  


# API functions
def refresh_agnitio_token(token_obj = None):
    # checks expiration of current token and fetches a fresh token if needed
    if token_obj:
        if token_obj['expiration'] > datetime.datetime.now():
            return token_obj, token_obj['token']
    url = "https://auth.emsicloud.com/connect/token"
    client_id = os.environ.get('API_EIS_ID')
    client_secret = os.environ.get('API_EIS_SECRET')

    payload = f"grant_type=client_credentials&client_id={client_id}&client_secret={client_secret}&scope=emsiauth"
    headers = {'content-type': 'application/x-www-form-urlencoded'}
    try:
        response = requests.request("POST", url, data=payload, headers=headers)
        new_token_object = {
            'expiration': datetime.datetime.now() + datetime.timedelta(seconds=3600),
            'token': response.json()['access_token']
        }
        return new_token_object, new_token_object['token']
    except:
        raise Exception(f'Token call status: {response.status_code}, token call response {response.text}')

def download_data(url, token_obj, payload=None):
    # makes api call and returns response. requires url, payload if post request, and token, which it updates if necessary and returns with the response
    fresh_token_obj, token = refresh_agnitio_token(token_obj)
    headers = {
        'content-type': "application/json",
        'authorization': f"Bearer {token}"
    }
    if payload==None:
        response = requests.get(url, headers=headers)
    else:
        response = requests.post(url, json=payload, headers=headers)
    if response.status_code != 200:
        print(url)
        print(token)
        print(payload)
        print(response.status_code)
        print(response.text)
    return response, fresh_token_obj

def get_agnitio_df(url, token, payload=None):
    # makes api calls and creates df from the response. requires url, payload if post request, and token, which it updates if necessary and returns with response
    response, token = download_data(url, token, payload)
    df = pd.DataFrame()
    if payload is None:
        # if a get request, return the hierarchy requested
        df = pd.DataFrame(response.json()['hierarchy'])
    else:
        # for each column in the response from agnitio, add that column to the dataframe,
        # and use the rows from the response as the data for the column
        try:
            for column in response.json()['data']:
                df[column['name']] = column['rows']
        except JSONDecodeError:
            print(url)
            print(payload)
            print(response.text)
            print(response.status_code)
            for column in response.json()['data']:
                df[column['name']] = column['rows']
    return df, token

def get_regional_completers(unitid, yearColumns, token, dataRun):
    url = f'https://agnitio.emsicloud.com/emsi.us.completers/{dataRun}/'
    payload = {
        "metrics": [{"name": f"Completers.{year}", "as": f"{year}"} for year in yearColumns],
        "constraints": [
            {
                "dimensionName": "Unit",
                "map": {str(unitid): [unitid]}
            },
            {
                "dimensionName": "Program",
                "asIdentity": True
            },
            {
                "dimensionName": "AwardLevel",
                "asIdentity": True
            }
        ],
        "zeroFill": False
    }

    df, token = get_agnitio_df(url, token, payload)

    df.rename(columns = {'Program': 'cip', 'Unit': 'unitid'}, inplace = True)

    metaUrl = 'https://agnitio.emsicloud.com/meta/dataset/emsi.us.completers/%s/Program' % dataRun
    programDf, token = get_agnitio_df(metaUrl, token)
    programDf.rename(columns = {'child': 'cip', 'name': 'Program Name'}, inplace = True)
    df = df.merge(programDf, on='cip', how='left')

    metaUrl = 'https://agnitio.emsicloud.com/meta/dataset/emsi.us.completers/%s/AwardLevel' % dataRun
    awardLevelDf, token = get_agnitio_df(metaUrl, token)
    awardLevelDf.rename(columns = {'child': 'AwardLevel', 'name': 'Award'}, inplace = True)
    df = df.merge(awardLevelDf, on='AwardLevel', how='left')
    df['TransferTrack'] = ''

    keepColumns = ['cip', 'Program Name', 'Award', 'TransferTrack']
    for year in yearColumns:
        keepColumns.append('%s' % str(year))

    df = df[keepColumns]

    return df   

def get_completers(unitid, dataRun, completionsYear):
    token_obj, token = refresh_agnitio_token()

    x = 0
    yearList = []
    while x < 3:
        yearList.append(int(completionsYear) - x)
        x = x + 1

    yearList.sort()

    regionalDf = get_regional_completers(unitid, yearList, token_obj, dataRun)

    regionalDf = regionalDf.rename(columns={'cip': "CIP Code", 'Program Name': "CIP Title", 'Award': 'Award Level', 'TransferTrack':'Transfer-track (yes or no)'})

    return regionalDf

def get_pdga_data_survey(ipeds, data_run, completions_year, save_location):
    template_location = r"G:\\Shared drives\\EDU Consulting - Private\\PDGA report masters and surveys\\US\\PDGA Survey Template\\PDGA_AutoSurveyTemplate.xlsx"
    wb = load_workbook(template_location)
    writer = pd.ExcelWriter(save_location, engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    regionCompletersDf = get_completers(ipeds, data_run, completions_year)

    regionCompletersDf.to_excel(writer, sheet_name='Credit Data', index=False, startcol=2, startrow=4, header=True)
                    # regionCipMapping.to_excel(writer, sheet_name='Program to Occupation Map', index=False, startcol=2, startrow=5, header=False)
    writer.save()
    return

if __name__ == '__main__':
    get_pdga_data_survey('U of I', 142285, '2022.4', '2021', r"G:\\Shared drives\\EDU Consulting - Private\\PDGA report masters and surveys\\US\\PDGA Survey Template")
